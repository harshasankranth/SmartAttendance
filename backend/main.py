import cv2
import pickle
import numpy as np
import base64
import json
import os
import shutil
from datetime import datetime
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List
from dotenv import load_dotenv
from supabase import create_client, Client
import warnings
warnings.filterwarnings('ignore')
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

app = FastAPI(title="SmartAttendance API", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"]
)

# ─────────────────────────────────────────────
# DATA MODELS
# ─────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str

class StudentRequest(BaseModel):
    name: str
    enrollment: str
    images: List[str]

# ─────────────────────────────────────────────
# STORAGE
# ─────────────────────────────────────────────
known_encodings = []
known_names = []
known_enrollments = []
students_db = {}
attendance_log = []

STUDENTS_FILE = "model/students.json"
ATTENDANCE_FILE = "model/attendance.json"
ENCODINGS_FILE = "model/face_encodings.pkl"

ADMIN_USERNAME = "harsha"
ADMIN_PASSWORD = "lucky123"

# ─────────────────────────────────────────────
# FACE RECOGNITION ENGINE
# ─────────────────────────────────────────────
import insightface
from insightface.app import FaceAnalysis

face_app = FaceAnalysis(name="buffalo_sc", providers=['CPUExecutionProvider'])
face_app.prepare(ctx_id=0, det_size=(640, 640))

def get_face_embedding(image):
    """Get face embedding from image using InsightFace"""
    try:
        faces = face_app.get(image)
        if len(faces) == 0:
            return None
        # Return embedding of largest face
        largest = max(faces, key=lambda f: (f.bbox[2]-f.bbox[0]) * (f.bbox[3]-f.bbox[1]))
        return largest.embedding, largest.bbox
    except:
        return None

def compare_faces(embedding, threshold=0.4):
    """Compare embedding against known faces"""
    if len(known_encodings) == 0:
        return "Unknown", 0, -1
    
    known_array = np.array(known_encodings)
    # Cosine similarity
    embedding_norm = embedding / np.linalg.norm(embedding)
    known_norm = known_array / np.linalg.norm(known_array, axis=1, keepdims=True)
    similarities = np.dot(known_norm, embedding_norm)
    
    best_idx = np.argmax(similarities)
    best_score = similarities[best_idx]
    
    if best_score >= threshold:
        return known_names[best_idx], float(best_score * 100), best_idx
    return "Unknown", float(best_score * 100), -1

# ─────────────────────────────────────────────
# SAVE / LOAD
# ─────────────────────────────────────────────
def save_encodings():
    os.makedirs("model", exist_ok=True)
    data = {
        "encodings": [e.tolist() for e in known_encodings],
        "names": known_names,
        "enrollments": known_enrollments
    }
    pickle.dump(data, open(ENCODINGS_FILE, "wb"))
    # Upload to Supabase
    try:
        with open(ENCODINGS_FILE, "rb") as f:
            supabase.storage.from_("models").upload(
                "face_encodings.pkl", f.read(),
                {"content-type": "application/octet-stream", "upsert": "true"}
            )
        print("Encodings uploaded to Supabase!")
    except Exception as e:
        print(f"Supabase upload error: {e}")

def load_encodings():
    global known_encodings, known_names, known_enrollments
    # Try local first
    if os.path.exists(ENCODINGS_FILE):
        try:
            data = pickle.load(open(ENCODINGS_FILE, "rb"))
            known_encodings = [np.array(e) for e in data["encodings"]]
            known_names = data["names"]
            known_enrollments = data["enrollments"]
            print(f"Loaded {len(known_names)} face encodings locally!")
            return True
        except:
            pass
    # Try Supabase
    try:
        print("Downloading encodings from Supabase...")
        os.makedirs("model", exist_ok=True)
        res = supabase.storage.from_("models").download("face_encodings.pkl")
        with open(ENCODINGS_FILE, "wb") as f:
            f.write(res)
        data = pickle.load(open(ENCODINGS_FILE, "rb"))
        known_encodings = [np.array(e) for e in data["encodings"]]
        known_names = data["names"]
        known_enrollments = data["enrollments"]
        print(f"Loaded {len(known_names)} encodings from Supabase!")
        return True
    except Exception as e:
        print(f"No encodings found: {e}")
        return False

def save_students():
    os.makedirs("model", exist_ok=True)
    json.dump(students_db, open(STUDENTS_FILE, "w"), indent=2)

def save_attendance():
    os.makedirs("model", exist_ok=True)
    json.dump(attendance_log, open(ATTENDANCE_FILE, "w"), indent=2)

# ─────────────────────────────────────────────
# STARTUP
# ─────────────────────────────────────────────
@app.on_event("startup")
async def startup():
    global students_db, attendance_log

    if os.path.exists(STUDENTS_FILE):
        students_db = json.load(open(STUDENTS_FILE))
        print(f"Loaded {len(students_db)} students")

    if os.path.exists(ATTENDANCE_FILE):
        attendance_log = json.load(open(ATTENDANCE_FILE))

    load_encodings()

# ─────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────
@app.post("/auth/login")
def login(req: LoginRequest):
    if req.username == ADMIN_USERNAME and req.password == ADMIN_PASSWORD:
        return {"success": True, "message": "Login successful"}
    raise HTTPException(status_code=401, detail="Invalid credentials")

# ─────────────────────────────────────────────
# STUDENTS
# ─────────────────────────────────────────────
@app.get("/students/all")
def get_students():
    return {"students": list(students_db.values()), "total": len(students_db)}

@app.post("/students/add")
async def add_student(req: StudentRequest):
    global students_db, known_encodings, known_names, known_enrollments

    name = req.name.strip()
    enrollment = req.enrollment.strip()

    if enrollment in students_db:
        raise HTTPException(status_code=400, detail="Enrollment already exists")

    folder = f"dataset/{enrollment}_{name.replace(' ', '_')}"
    os.makedirs(folder, exist_ok=True)

    saved = 0
    new_encodings = []

    for i, img_b64 in enumerate(req.images):
        try:
            img_bytes = base64.b64decode(img_b64)
            nparr = np.frombuffer(img_bytes, np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            if frame is None:
                continue

            result = get_face_embedding(frame)
            if result is None:
                continue

            embedding, bbox = result
            new_encodings.append(embedding)
            cv2.imwrite(f"{folder}/img_{i:03d}.jpg", frame)
            saved += 1
        except:
            continue

    if len(new_encodings) < 3:
        shutil.rmtree(folder)
        raise HTTPException(status_code=400, detail="Could not detect face in enough photos. Please ensure good lighting and face the camera directly.")

    # Add to known faces
    for enc in new_encodings:
        known_encodings.append(enc)
        known_names.append(f"{enrollment}_{name.replace(' ', '_')}")
        known_enrollments.append(enrollment)

    students_db[enrollment] = {
        "name": name,
        "enrollment": enrollment,
        "folder": folder,
        "photos": saved,
        "encodings": len(new_encodings),
        "added_on": datetime.now().strftime("%d %b %Y %H:%M")
    }

    save_students()
    save_encodings()

    # Save to Supabase
    try:
        supabase.table("students").insert({
            "name": name,
            "enrollment": enrollment,
            "folder": folder,
            "photos": saved,
            "added_on": students_db[enrollment]["added_on"]
        }).execute()
    except Exception as e:
        print(f"Supabase insert error: {e}")

    return {
        "success": True,
        "message": f"Student {name} added with {len(new_encodings)} face encodings",
        "trained": True
    }

@app.delete("/students/delete/{enrollment}")
async def delete_student(enrollment: str):
    global students_db, known_encodings, known_names, known_enrollments

    if enrollment not in students_db:
        raise HTTPException(status_code=404, detail="Student not found")

    student = students_db[enrollment]
    folder = student.get("folder", "")
    if folder and os.path.exists(folder):
        shutil.rmtree(folder)

    # Remove encodings
    label = f"{enrollment}_{student['name'].replace(' ', '_')}"
    indices = [i for i, n in enumerate(known_names) if n == label]
    known_encodings = [e for i, e in enumerate(known_encodings) if i not in indices]
    known_names = [n for i, n in enumerate(known_names) if i not in indices]
    known_enrollments = [e for i, e in enumerate(known_enrollments) if i not in indices]

    del students_db[enrollment]
    save_students()
    save_encodings()

    return {"success": True, "message": f"Student {student['name']} deleted"}

# ─────────────────────────────────────────────
# ATTENDANCE
# ─────────────────────────────────────────────
@app.get("/attendance")
def get_attendance():
    today = datetime.now().strftime("%d %b %Y")
    todays = [r for r in attendance_log if r["date"] == today]
    return {
        "date": today,
        "records": todays,
        "total_present": len(todays),
        "total_students": len(students_db)
    }

@app.get("/attendance/export")
def export_attendance():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        header_fill = PatternFill("solid", fgColor="1E40AF")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["#", "Name", "Enrollment", "Date", "Time", "Confidence"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        today = datetime.now().strftime("%d %b %Y")
        todays = [r for r in attendance_log if r["date"] == today]

        for row, record in enumerate(todays, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=record.get("name", ""))
            ws.cell(row=row, column=3, value=record.get("enrollment", ""))
            ws.cell(row=row, column=4, value=record.get("date", ""))
            ws.cell(row=row, column=5, value=record.get("time", ""))
            ws.cell(row=row, column=6, value=record.get("confidence", ""))

        for col in ["A","B","C","D","E","F"]:
            ws.column_dimensions[col].width = 20

        export_path = "model/attendance_export.xlsx"
        wb.save(export_path)
        return FileResponse(
            export_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"attendance_{today}.xlsx"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ─────────────────────────────────────────────
# WEBSOCKET — LIVE RECOGNITION
# ─────────────────────────────────────────────
@app.websocket("/ws/recognize")
async def recognize(websocket: WebSocket):
    await websocket.accept()
    print("Frontend connected!")
    marked_this_session = set()

    try:
        while True:
            data = await websocket.receive_text()
            payload = json.loads(data)
            image_b64 = payload.get("image", "")

            if not image_b64 or len(known_encodings) == 0:
                await websocket.send_json({
                    "faces": [],
                    "marked_today": list(marked_this_session),
                    "total_present": len(marked_this_session)
                })
                continue

            img_bytes = base64.b64decode(image_b64)
            nparr = np.frombuffer(img_bytes, np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

            if frame is None:
                continue

            results = []

            try:
                faces = face_app.get(frame)

                for face in faces:
                    embedding = face.embedding
                    bbox = face.bbox.astype(int)
                    x, y, x2, y2 = bbox
                    w, h = x2-x, y2-y

                    name, confidence, idx = compare_faces(embedding, threshold=0.35)

                    if name != "Unknown":
                        # Get clean name and enrollment
                        parts = name.split("_", 1)
                        enrollment = parts[0] if len(parts) > 1 else "N/A"
                        student_info = students_db.get(enrollment, {})
                        display_name = student_info.get("name", name)

                        if confidence >= 60 and name not in marked_this_session:
                            marked_this_session.add(name)
                            record = {
                                "name": display_name,
                                "enrollment": enrollment,
                                "confidence": round(confidence, 2),
                                "time": datetime.now().strftime("%H:%M:%S"),
                                "date": datetime.now().strftime("%d %b %Y")
                            }
                            attendance_log.append(record)
                            save_attendance()
                            try:
                                supabase.table("attendance").insert(record).execute()
                            except:
                                pass
                            print(f"MARKED: {display_name} {confidence:.1f}%")

                        results.append({
                            "name": display_name if confidence >= 55 else "Unknown",
                            "enrollment": enrollment if confidence >= 55 else "",
                            "confidence": round(confidence, 2),
                            "box": {"x": int(x), "y": int(y), "w": int(w), "h": int(h)}
                        })
                    else:
                        results.append({
                            "name": "Unknown",
                            "enrollment": "",
                            "confidence": round(confidence, 2),
                            "box": {"x": int(x), "y": int(y), "w": int(w), "h": int(h)}
                        })

            except Exception as e:
                print(f"Recognition error: {e}")

            await websocket.send_json({
                "faces": results,
                "marked_today": list(marked_this_session),
                "total_present": len(marked_this_session)
            })

    except WebSocketDisconnect:
        print("Frontend disconnected.")

# ─────────────────────────────────────────────
# HEALTH CHECK
# ─────────────────────────────────────────────
@app.get("/")
def home():
    return {
        "status": "running",
        "students": len(students_db),
        "model_loaded": len(known_encodings) > 0,
        "total_encodings": len(known_encodings)
    }